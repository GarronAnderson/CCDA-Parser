import openpyxl
import sqlite3 as sqlite

wb = openpyxl.load_workbook(r"Raw Data\CDC_REC.xlsx")

sheet = wb["All Codes"]

codes = {}

for row in range(2, 968):
    k = sheet.cell(row=row, column=1)
    v = sheet.cell(row=row, column=3)
    codes.update({k.value: v.value})
